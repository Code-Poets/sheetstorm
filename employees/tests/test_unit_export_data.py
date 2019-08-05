import csv
import io
import zipfile

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from freezegun import freeze_time
from openpyxl import load_workbook

from employees.common.constants import ExcelGeneratorSettingsConstants as excel_constants
from employees.common.exports import generate_xlsx_for_project
from employees.common.exports import generate_xlsx_for_single_user
from employees.common.exports import get_employee_name
from employees.common.exports import save_work_book_as_csv
from employees.common.exports import save_work_book_as_zip_of_csv
from employees.factories import ReportFactory
from employees.models import Report
from managers.factories import ProjectFactory
from users.factories import AdminUserFactory
from users.factories import ManagerUserFactory
from users.factories import UserFactory


@freeze_time("2019-06-24")
class DataSetUpToTests(TestCase):
    def setUp(self):
        super().setUp()
        self.user = AdminUserFactory()
        self.project_start_date = timezone.now() + timezone.timedelta(days=-1)
        self.project = ProjectFactory(
            name="aaa", start_date=self.project_start_date, stop_date=timezone.now() + timezone.timedelta(days=6)
        )
        self.report = ReportFactory(author=self.user, project=self.project)
        self.project.members.add(self.user)
        self.data = {"project": self.project.pk, "author": self.user.pk}
        self.url_single_user = reverse(
            "export-data",
            kwargs={
                "pk": self.user.pk,
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )
        self.url_project = reverse(
            "export-project-reports",
            kwargs={
                "pk": self.data["project"],
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )
        self.workbook_for_project = generate_xlsx_for_project(self.project)
        self.workbook_for_user = generate_xlsx_for_single_user(self.user)


class ExportViewTest(DataSetUpToTests):
    def test_export_reports_should_download_for_single_user_if_he_is_logged(self):
        self.client.force_login(self.user)
        response = self.client.get(self.url_single_user)
        self.assertEqual(response.status_code, 200)

    def test_export_reports_should_not_download_data_if_he_is_not_logged(self):
        response = self.client.get(self.url_single_user)
        self.assertEqual(response.status_code, 302)

    def test_export_reports_for_project_should_download_if_user_is_logged(self):
        self.client.force_login(self.user)
        response = self.client.get(self.url_project)
        self.assertEqual(response.status_code, 200)

    def test_export_reports_for_project_should_not_download_if_user_is_not_logged(self):
        response = self.client.get(self.url_project)
        self.assertEqual(response.status_code, 302)

    def test_export_reports_should_download_for_single_user_csv(self):
        self.client.force_login(self.user)
        response = self.client.get(self.url_single_user + "?format=csv")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response._headers["content-disposition"][1].endswith('csv"'))

    def test_export_reports_for_project_should_download_for_project_csv_as_zip_file(self):
        self.client.force_login(self.user)
        response = self.client.get(self.url_project + "?format=csv")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response._headers["content-disposition"][1].endswith('zip"'))

    def test_export_reports_for_project_author_reports_should_download_if_user_is_manager_of_current_project(self):
        manager = ManagerUserFactory()
        self.project.managers.add(manager)
        self.client.force_login(manager)
        project_author_reports_url = reverse(
            "export-project-author-reports",
            kwargs={
                "pk": self.data["project"],
                "user_pk": self.user.pk,
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )
        response = self.client.get(project_author_reports_url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response._headers["content-disposition"][1].endswith('xlsx"'))

    def test_export_reports_for_project_author_reports_should_download_csv(self):
        manager = ManagerUserFactory()
        self.project.managers.add(manager)
        self.client.force_login(manager)
        project_author_reports_url = reverse(
            "export-project-author-reports",
            kwargs={
                "pk": self.data["project"],
                "user_pk": self.user.pk,
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )

        response = self.client.get(project_author_reports_url + "?format=csv")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response._headers["content-disposition"][1].endswith('csv"'))

    def test_export_reports_for_project_author_reports_should_download_xlsx_if_user_is_not_a_manager_of_current_project(
        self
    ):
        manager = ManagerUserFactory()
        self.client.force_login(manager)
        project_author_reports_url = reverse(
            "export-project-author-reports",
            kwargs={
                "pk": self.data["project"],
                "user_pk": self.user.pk,
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )

        response = self.client.get(project_author_reports_url)
        self.assertEqual(response.status_code, 404)

    def test_export_reports_for_project_author_reports_should_download_csv_if_user_is_not_a_manager_of_current_project(
        self
    ):
        manager = ManagerUserFactory()
        self.client.force_login(manager)
        project_author_reports_url = reverse(
            "export-project-author-reports",
            kwargs={
                "pk": self.data["project"],
                "user_pk": self.user.pk,
                "year": str(self.report.creation_date.year),
                "month": str(self.report.creation_date.month),
            },
        )

        response = self.client.get(project_author_reports_url + "?format=csv")

        self.assertEqual(response.status_code, 404)


class ExportMethodTestForProject(DataSetUpToTests):
    def test_sheetnames_should_have_name_and_first_letter_of_surname_and_one_sheet(self):
        authors = self.workbook_for_project.sheetnames
        for author in authors:
            self.assertEqual(author, f"{self.user.first_name} {self.user.last_name[0]}.")
        self.assertEqual(len(authors), 1)

    def test_date_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.date,
            str(
                self.workbook_for_project.active.cell(
                    row=excel_constants.FIRST_ROW_FOR_DATA.value,
                    column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value[
                        excel_constants.DATE_HEADER_STR.value
                    ].position,
                ).value
            ),
        )

    def test_task_activity_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.task_activities.name,
            self.workbook_for_project.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value[
                    excel_constants.TASK_ACTIVITY_HEADER_STR.value
                ].position,
            ).value,
        )

    def test_hours_should_be_the_same_in_excel(self):
        work_hours = f"{self.report.work_hours_str}"
        self.assertEqual(
            excel_constants.TIMEVALUE_FORMULA.value.format(work_hours),
            self.workbook_for_project.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value[
                    excel_constants.HOURS_HEADER_STR.value
                ].position,
            ).value,
        )

    def test_description_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.description,
            self.workbook_for_project.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value[
                    excel_constants.DESCRIPTION_HEADER_STR.value
                ].position,
            ).value,
        )

    @staticmethod
    def _hours_date_time_to_excel_time_field(hours_delta):
        sum_from_reports_as_string = f"{int(hours_delta.total_seconds() / 3600)}:00:00"
        return f'=timevalue("{sum_from_reports_as_string}")'


class ExportMethodTestForSingleUser(DataSetUpToTests):
    def test_sheetnames_should_have_name_and_first_letter_of_surname_and_one_sheet(self):
        sheet_names = self.workbook_for_user.sheetnames
        for author in sheet_names:
            self.assertEqual(author, f"{self.user.first_name} {self.user.last_name[0]}.")
        self.assertEqual(len(sheet_names), 1)

    def test_date_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.date,
            str(
                self.workbook_for_user.active.cell(
                    row=excel_constants.FIRST_ROW_FOR_DATA.value,
                    column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                        excel_constants.DATE_HEADER_STR.value
                    ].position,
                ).value
            ),
        )

    def test_project_name_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.project.name,
            str(
                self.workbook_for_user.active.cell(
                    row=excel_constants.FIRST_ROW_FOR_DATA.value,
                    column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                        excel_constants.PROJECT_HEADER_STR.value
                    ].position,
                ).value
            ),
        )

    def test_task_activity_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.task_activities.name,
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.TASK_ACTIVITY_HEADER_STR.value
                ].position,
            ).value,
        )

    def test_hours_should_be_the_same_in_excel(self):
        work_hours = f"{self.report.work_hours_str}"
        self.assertEqual(
            excel_constants.TIMEVALUE_FORMULA.value.format(work_hours),
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.HOURS_HEADER_STR.value
                ].position,
            ).value,
        )

    def test_description_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.report.description,
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.DESCRIPTION_HEADER_STR.value
                ].position,
            ).value,
        )


class SaveWorkBookAsCSVTesCase(DataSetUpToTests):
    def setUp(self):
        super().setUp()
        self.csv_file = io.StringIO()
        writer = csv.writer(self.csv_file)
        hours_column_setting = excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
            excel_constants.HOURS_HEADER_STR.value
        ]
        save_work_book_as_csv(writer, self.workbook_for_user, hours_column_setting)
        self.csv_content = [line for line in csv.reader(self.csv_file.getvalue().split("\n"))]

    def test_date_should_be_the_same_in_excel(self):
        self.assertEqual(
            str(
                self.workbook_for_user.active.cell(
                    row=excel_constants.FIRST_ROW_FOR_DATA.value,
                    column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                        excel_constants.DATE_HEADER_STR.value
                    ].position,
                ).value
            ),
            self.csv_content[excel_constants.FIRST_ROW_FOR_DATA.value - 2][0],
        )

    def test_project_name_should_be_the_same_in_excel(self):
        self.assertEqual(
            str(
                self.workbook_for_user.active.cell(
                    row=excel_constants.FIRST_ROW_FOR_DATA.value,
                    column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                        excel_constants.PROJECT_HEADER_STR.value
                    ].position,
                ).value
            ),
            self.csv_content[excel_constants.FIRST_ROW_FOR_DATA.value - 2][1],
        )

    def test_task_activity_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.TASK_ACTIVITY_HEADER_STR.value
                ].position,
            ).value,
            self.csv_content[excel_constants.FIRST_ROW_FOR_DATA.value - 2][2],
        )

    def test_hours_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.HOURS_HEADER_STR.value
                ].position,
            ).value,
            f'=timevalue("{self.csv_content[excel_constants.FIRST_ROW_FOR_DATA.value - 2][3]}")',
        )

    def test_description_should_be_the_same_in_excel(self):
        self.assertEqual(
            self.workbook_for_user.active.cell(
                row=excel_constants.FIRST_ROW_FOR_DATA.value,
                column=excel_constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value[
                    excel_constants.DESCRIPTION_HEADER_STR.value
                ].position,
            ).value,
            self.csv_content[excel_constants.FIRST_ROW_FOR_DATA.value - 2][4],
        )


class SaveWorkBookAsZipOfCSVTesCase(TestCase):
    def setUp(self):
        super().setUp()
        self.user = UserFactory()
        self.project = ProjectFactory()
        self.project.members.add(self.user)
        ReportFactory(author=self.user, project=self.project)

    def test_save_work_book_as_zip_of_csv_creates_zip_file(self):
        workbook_for_project = generate_xlsx_for_project(self.project)
        zip_file_data = save_work_book_as_zip_of_csv(workbook_for_project)

        self.assertIsInstance(zip_file_data, io.BytesIO)

        zip_file = zipfile.ZipFile(zip_file_data)
        self.assertEqual(len(zip_file.infolist()), 1)
        self.assertEqual(
            zip_file.infolist()[0].filename,
            f"{self.user.first_name.lower()}_{self.user.last_name[0].lower()}-reports.csv",
        )
        self.assertTrue(zip_file.infolist()[0].file_size > 0)

    def test_save_work_book_as_zip_of_csv_creates_zip_file_with_two_csv_files_for_two_users_in_project(self):
        user_2 = UserFactory()
        self.project.members.add(user_2)
        ReportFactory(author=user_2, project=self.project)
        workbook_for_project = generate_xlsx_for_project(self.project)

        zip_file_data = save_work_book_as_zip_of_csv(workbook_for_project)
        self.assertIsInstance(zip_file_data, io.BytesIO)

        zip_file = zipfile.ZipFile(zip_file_data)
        self.assertEqual(len(zip_file.infolist()), 2)
        file_names = [zip_info.filename for zip_info in zip_file.infolist()]

        self.assertIn(f"{self.user.first_name.lower()}_{self.user.last_name[0].lower()}-reports.csv", file_names)
        self.assertIn(f"{user_2.first_name.lower()}_{user_2.last_name[0].lower()}-reports.csv", file_names)
        self.assertTrue(zip_file.infolist()[0].file_size > 0)
        self.assertTrue(zip_file.infolist()[1].file_size > 0)


class TestExportingFunctions(TestCase):
    def setUp(self):
        super().setUp()
        self.employee1 = UserFactory(first_name="Cezar", last_name="Goldstein")
        self.employee2 = UserFactory(first_name="", last_name="", email="bozydar.stolzman@codepots.it")
        self.employee3 = UserFactory(first_name="Abimelek", last_name="Zuckerberg")
        self.manager = ManagerUserFactory()
        self.project = ProjectFactory()
        self.project.members.add(self.employee1)
        self.project.members.add(self.employee2)
        self.project.members.add(self.employee3)
        reports_in_day = 2
        # creating reports in desc order
        number_of_days = 4
        self.year = "2019"
        self.month = "06"
        for i in range(number_of_days, 0, -1):
            for _ in range(reports_in_day):
                ReportFactory(author=self.employee2, project=self.project, date=f"{self.year}-{self.month}-{i}")
                ReportFactory(author=self.employee3, project=self.project, date=f"{self.year}-{self.month}-{i}")
                ReportFactory(author=self.employee1, project=self.project, date=f"{self.year}-{self.month}-{i}")
        self.report_asc = Report.objects.filter(author__id=self.employee1.pk).order_by("date")
        self.reports_per_user = reports_in_day * number_of_days

    def test_unsorted_reports_will_be_sorted_asc_by_first_name_and_asc_by_date_in_project_export(self):
        project_workbook = generate_xlsx_for_project(self.project)
        for j, sheet in enumerate(project_workbook.worksheets):
            # Check that sheets are sorted in ascending order, according to employee's first name / user email
            self.assertEqual(sheet.title, get_employee_name(getattr(self, f"employee{3 - j}")))
            self._assert_reports_are_sorted_in_ascending_order(project_workbook)
            self._assert_dates_are_unique_in_reports_of_user(project_workbook)

    def test_project_members_with_no_report_will_be_skipped_in_project_export(self):
        not_a_member = UserFactory(first_name="Asterix", last_name="Longsword")
        self.project.members.add(not_a_member)
        project_workbook = generate_xlsx_for_project(self.project)
        employee_name = get_employee_name(not_a_member)
        self.assertNotIn(employee_name, [s.title for s in project_workbook.worksheets])

    def test_unsorted_reports_will_be_sorted_asc_by_date_in_user_export(self):
        user_workbook = generate_xlsx_for_single_user(self.employee1)
        self._assert_reports_are_sorted_in_ascending_order(user_workbook)
        self._assert_dates_are_unique_in_reports_of_user(user_workbook)

    def _assert_reports_are_sorted_in_ascending_order(self, workbook):
        for i, element in enumerate(self.report_asc):
            # there are two reports per day, but "Date" should occur only once per day
            if i % 2 == 0:
                self.assertEqual(
                    workbook.active.cell(row=excel_constants.FIRST_ROW_FOR_DATA.value + i, column=1).value, element.date
                )
            else:
                self.assertEqual(
                    workbook.active.cell(row=excel_constants.FIRST_ROW_FOR_DATA.value + i, column=1).value, None
                )

    def _assert_dates_are_unique_in_reports_of_user(self, workbook):
        for worksheet in workbook.worksheets:
            dates = []
            for i in range(self.reports_per_user):
                cell_value = worksheet.cell(row=excel_constants.FIRST_ROW_FOR_DATA.value + i, column=1).value
                if cell_value is not None:
                    dates.append(cell_value)
            number_of_dates = len(dates)
            number_of_unique_dates = len(set(dates))
            self.assertTrue(number_of_dates > 0)
            self.assertEqual(number_of_dates, number_of_unique_dates)

    def test_user_can_export_only_his_own_reports(self):
        self.client.force_login(self.employee1)
        url = reverse("export-data", kwargs={"pk": self.employee2.pk, "year": self.year, "month": self.month})
        response = self.client.get(url)
        received_workbook = load_workbook(filename=io.BytesIO(response.content))
        self.assertEqual(len(received_workbook.sheetnames), 1)
        self.assertEqual(received_workbook.sheetnames[0], f"{self.employee1.first_name} {self.employee1.last_name[0]}.")
