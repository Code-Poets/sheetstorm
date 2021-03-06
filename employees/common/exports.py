import re
from datetime import timedelta
from typing import Dict
from typing import List
from typing import Optional

from django.utils.datetime_safe import datetime
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import _writer

from common.convert import timedelta_to_string
from employees.common.constants import ColumnSettings
from employees.common.constants import ExcelGeneratorSettingsConstants as constants
from employees.models import Report
from employees.models import ReportQuerySet
from managers.models import Project
from users.models import CustomUser


def set_format_styles_for_main_cells(cell: Cell, is_header: bool) -> None:
    cell.font = Font(name=constants.FONT.value, bold=True)
    cell.alignment = Alignment(horizontal=constants.CENTER_ALINGMENT.value)
    cell.border = (
        Border(
            bottom=constants.BORDER_STYLE.value,
            top=constants.BORDER_STYLE.value,
            right=constants.BORDER_STYLE.value,
            left=constants.BORDER_STYLE.value,
        )
        if is_header
        else Border(top=constants.BORDER_STYLE.value)
    )


def set_and_fill_cell(cell: Cell, cell_value: str) -> None:
    wrapped_alignment = Alignment(vertical=constants.VERCTICAL_TOP.value, wrap_text=True)
    cell.alignment = wrapped_alignment
    cell.value = cell_value
    set_borders_between_columns(cell)


def set_and_fill_hours_cell(cell: Cell, cell_value: str) -> None:
    if cell_value is not None:
        cell.value = cell_value
    else:
        cell.value = cell_value
    alignment = Alignment(vertical=constants.VERCTICAL_TOP.value, horizontal=constants.CENTER_ALINGMENT.value)
    cell.alignment = alignment
    cell.number_format = constants.HOURS_FORMAT.value
    set_borders_between_columns(cell)


def set_borders_between_columns(cell: Cell) -> None:
    cell.border = Border(left=constants.BORDER_STYLE.value, right=constants.BORDER_STYLE.value)


def separate_days(cell: Cell) -> None:
    cell.border = Border(
        left=constants.BORDER_STYLE.value, right=constants.BORDER_STYLE.value, top=constants.BORDER_STYLE.value
    )


def get_employee_name(author: CustomUser) -> str:
    if author.last_name and author.first_name:
        return f"{author.first_name} {author.last_name[0]}."
    else:
        return f"{author.email}"


def generate_xlsx_for_single_user(author: CustomUser) -> Workbook:
    return ReportExtractor().generate_xlsx_for_single_user(author)


def generate_xlsx_for_project(project: Project) -> Workbook:
    return ReportExtractor().generate_xlsx_for_project(project)


class ReportExtractor:
    def __init__(self) -> None:
        self._current_row = -1
        self._workbook = None  # type: Workbook
        self._headers = []  # type: List
        self._description_column_index = -1
        self._hours_column_index = -1
        self._daily_hours_column_index = -1
        self._sum_hours = None  # type: datetime.timedelta
        self._last_date = None
        self._active_worksheet = None  # type: Workbook
        self._headers_settings = {}  # type: Dict

    def generate_xlsx_for_project(self, project: Project) -> Workbook:
        authors = project.members.all()
        self._workbook = Workbook()
        self._set_xlsx_settings_for_project_report()
        for author in authors:
            employee_name = get_employee_name(author)
            reports = project.report_set.filter(author=author.pk).order_by("date")
            if not reports:
                continue

            self._fill_report_for_single_user(employee_name, reports)
            self._set_printing_settings_for_current_sheet()

        self._workbook._sheets.sort(key=lambda w: str.lower(w.title))
        return self._workbook

    def generate_xlsx_for_single_user(self, author: CustomUser) -> Workbook:
        reports = author.get_reports_created().order_by("date", "project__name")  # type: ReportQuerySet
        self._workbook = Workbook()
        self._set_xlsx_settings_for_user_report()
        employee_name = get_employee_name(author)

        self._fill_report_for_single_user(employee_name, reports)
        self._set_printing_settings_for_current_sheet()
        return self._workbook

    def _fill_report_for_single_user(self, employee_name: str, reports: ReportQuerySet) -> None:
        self._prepare_worksheet(employee_name)
        for report in reports:
            self._fill_single_report(report)

        self._summarize_user_reports()

    def _fill_single_report(self, report: Report) -> None:
        report_date = self._get_report_date(report)
        report_description = self.delete_illigal_characters(report.description)
        storage_data = {
            constants.DATE_HEADER_STR.value: report_date,
            constants.PROJECT_HEADER_STR.value: report.project.name,
            constants.TASK_ACTIVITY_HEADER_STR.value: report.task_activities.name,
            constants.HOURS_HEADER_STR.value: report.work_hours,
            constants.DESCRIPTION_HEADER_STR.value: report_description,
        }
        self._fill_current_report_data(storage_data)
        self._set_row_height(str(storage_data[constants.DESCRIPTION_HEADER_STR.value]))
        self._current_row += 1

    def _prepare_worksheet(self, employee_name: str) -> None:
        self._reset_per_sheet_settings()
        self._set_active_worksheet_name(employee_name)
        self._fill_headers(employee_name)

    def _reset_per_sheet_settings(self) -> None:
        self._current_row = constants.FIRST_ROW_FOR_DATA.value
        self._last_date = None
        self._sum_hours = None

    def _set_xlsx_settings_for_project_report(self) -> None:
        self._headers_settings = dict(constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value)
        self._headers = [k for k, v in self._headers_settings.items() if v is not None]

    def _set_xlsx_settings_for_user_report(self) -> None:
        self._headers_settings = dict(constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_SINGLE_USER.value)
        self._headers = [k for k, v in self._headers_settings.items() if v is not None]

    def _set_active_worksheet_name(self, employee_name: str) -> None:
        if self._workbook.active.title != "Sheet":
            self._set_employees_worksheet_active(employee_name)
        self._active_worksheet = self._workbook.active
        self._active_worksheet.title = employee_name

    def _set_employees_worksheet_active(self, employee_name: str) -> None:
        try:
            employee_sheet_index = self._workbook.worksheets.index(employee_name)
        except ValueError:
            self._workbook.create_sheet()
            employee_sheet_index = -1
        self._workbook.active = self._workbook.worksheets[employee_sheet_index]

    def _fill_headers(self, employee_name: str) -> None:
        self._set_employee_name_in_worksheet(employee_name)
        for col_num, column_name in enumerate(self._headers, start=1):
            cell = self._active_worksheet.cell(row=constants.HEADERS_ROW.value, column=col_num)
            cell.value = column_name
            set_format_styles_for_main_cells(cell, is_header=True)
            self._set_column_width(col_num, column_name)

    def _set_employee_name_in_worksheet(self, employee_name: str) -> None:
        self._active_worksheet.merge_cells(
            start_row=constants.EMPLOYEE_NAME_ROW.value,
            end_row=constants.EMPLOYEE_NAME_ROW.value,
            start_column=constants.EMPLOYEE_NAME_START_COLUMN.value,
            end_column=constants.EMPLOYEE_NAME_END_COLUMN.value,
        )
        cell = self._active_worksheet.cell(
            row=constants.EMPLOYEE_NAME_ROW.value, column=constants.EMPLOYEE_NAME_START_COLUMN.value
        )
        cell.value = constants.EMPLOYEE_NAME.value.format(employee_name)

    def _set_column_width(self, col_num: int, column_name: str) -> None:
        column_letter = get_column_letter(col_num)
        column_dimensions = self._active_worksheet.column_dimensions[column_letter]
        column_dimensions.width = self._headers_settings[column_name].width

    def _fill_current_report_data(self, storage_data: dict) -> None:
        is_next_day = storage_data[constants.DATE_HEADER_STR.value] is not None

        for column_name, cell_value in storage_data.items():
            if self._headers_settings.get(column_name) is not None:
                cell = self._active_worksheet.cell(
                    row=self._current_row, column=self._headers_settings[column_name].position
                )
                if column_name == constants.HOURS_HEADER_STR.value:
                    set_and_fill_hours_cell(cell, cell_value)
                    self._sum_hours = self._sum_hours + cell_value if self._sum_hours is not None else cell_value
                else:
                    set_and_fill_cell(cell, cell_value)

                if is_next_day:
                    separate_days(cell)

    def _summarize_user_reports(self) -> None:
        total_cell = self._active_worksheet.cell(row=self._current_row, column=constants.TOTAL_COLUMN.value)
        total_cell.value = constants.TOTAL.value
        for column_number in range(1, self._active_worksheet.max_column + 1):
            set_format_styles_for_main_cells(
                self._active_worksheet.cell(row=self._current_row, column=column_number), is_header=False
            )

        total_hours_cell = self._active_worksheet.cell(
            row=self._current_row, column=self._headers_settings[constants.HOURS_HEADER_STR.value].position
        )
        total_hours_cell.value = self._sum_hours
        total_hours_cell.number_format = constants.TOTAL_HOURS_FORMAT.value
        set_format_styles_for_main_cells(total_hours_cell, is_header=False)

    def _get_report_date(self, current_report: Report) -> Optional[datetime]:
        date = current_report.date
        if self._last_date == date:
            report_date = None
        else:
            report_date = date
        self._last_date = date
        return report_date

    def _set_row_height(self, description: str) -> None:
        """
        If the description contains a new line ( "\n" character) or text exceeds 100 characters row height is automatically resized.
        """

        new_lines_in_description = description.count("\n")
        splitted_description = description.split("\n")
        rows_in_xlsx = 0

        for row in splitted_description:
            rows_in_xlsx += int(len(row) / constants.MAX_DESCRIPTION_WIDTH.value)

        row_height = constants.DEFAULT_ROW_HEIGHT.value * (1 + new_lines_in_description + rows_in_xlsx)
        self._active_worksheet.row_dimensions[self._current_row].height = row_height

    def _set_printing_settings_for_current_sheet(self) -> None:
        self._active_worksheet.set_printer_settings(paper_size=9, orientation="landscape")
        self._active_worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        self._active_worksheet.page_setup.fitToHeight = False

    def delete_illigal_characters(self, description: str) -> str:  # pylint: disable=no-self-use
        return re.sub(r"[\000-\010]|[\013-\014]|[\016-\037]", "", description)


def save_work_book_as_csv(writer: _writer, work_book: Workbook, hours_column_setting: ColumnSettings) -> None:
    sheet = work_book.active
    is_last_row = False

    for row_number, row in enumerate(sheet.rows, start=1):
        # Check if is last row by comparing first cell value to `TOTAL`.
        if isinstance(row[0].value, str) and row[0].value == constants.TOTAL.value:
            is_last_row = True

        next_row = []
        for cell_number, cell in enumerate(row, start=1):
            if row_number == 1:
                next_row.append(cell.parent.title)
                break
            if isinstance(cell.value, timedelta) and not is_last_row and cell_number == hours_column_setting.position:
                next_row.append(cell.value)
            elif isinstance(cell.value, timedelta) and is_last_row:
                next_row.append(timedelta_to_string(cell.value))
            else:
                next_row.append(cell.value)

        writer.writerow(next_row)


def export_all_project_reports_as_one_csv_file(work_book: Workbook, writer: _writer) -> None:
    hours_column_setting: ColumnSettings = constants.HEADERS_TO_COLUMNS_SETTINGS_FOR_USER_IN_PROJECT.value[
        constants.HOURS_HEADER_STR.value
    ]
    for sheet_index in range(len(work_book.sheetnames)):
        work_book.active = sheet_index
        save_work_book_as_csv(writer, work_book, hours_column_setting)
